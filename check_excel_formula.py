import openpyxl

def check_formula():
    try:
        wb = openpyxl.load_workbook('fcffsimpleginzu.xlsx', data_only=False)
        sheet = wb['Valuation output']
        
        print("Searching for 'Value of Equity' in 'Valuation output' sheet...")
        
        found = False
        for row in sheet.iter_rows(max_row=50, max_col=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'Value of Equity' in cell.value:
                    print(f"Found label at {cell.coordinate}: {cell.value}")
                    # The value is usually in the next column or nearby
                    # Let's check neighbors
                    r, c = cell.row, cell.column
                    neighbor = sheet.cell(row=r, column=c+1)
                    print(f"Neighbor ({neighbor.coordinate}) value/formula: {neighbor.value}")
                    
                    # Check for share price too
                    found = True
                
                if cell.value and isinstance(cell.value, str) and 'Value per share' in cell.value:
                     print(f"Found 'Value per share' at {cell.coordinate}: {cell.value}")
                     r, c = cell.row, cell.column
                     neighbor = sheet.cell(row=r, column=c+1)
                     print(f"Neighbor ({neighbor.coordinate}) value/formula: {neighbor.value}")

        if not found:
            print("Could not find 'Value of Equity' label.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_formula()
