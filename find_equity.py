import openpyxl

def find_equity():
    try:
        wb = openpyxl.load_workbook('fcffsimpleginzu.xlsx', data_only=False)
        sheet = wb['Valuation output']
        
        print("--- Searching for 'Equity' ---")
        for row in sheet.iter_rows(max_row=100, max_col=5):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'Equity' in cell.value:
                    print(f"Found '{cell.value}' at {cell.coordinate}")
                    # Print neighbors
                    r, c = cell.row, cell.column
                    neighbor = sheet.cell(row=r, column=c+1)
                    print(f"  Neighbor ({neighbor.coordinate}): {neighbor.value}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    find_equity()
